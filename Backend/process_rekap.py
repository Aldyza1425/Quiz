import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter

def process_rekap(input_csv, output_xlsx):
    # 1. Load Data
    try:
        df = pd.read_csv(input_csv)
    except FileNotFoundError:
        # Create dummy data if file not found for demonstration
        data = {
            'Nama Siswa': ['Sama Siswa', 'Budi Santoso', 'Siti Aminah', 'Andi Wijaya'],
            'Hadir': [0, 20, 18, 15],
            'Izin': [1, 0, 1, 2],
            'Sakit': [0, 0, 1, 0],
            'Alpha': [0, 0, 0, 3],
            'Persentase %': ['0%', '100%', '90%', '75%']
        }
        df = pd.DataFrame(data)
        df.to_csv(input_csv, index=False)
        print(f"File {input_csv} tidak ditemukan, membuat file dummy untuk demonstrasi.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Absensi Kelas XII IPA"

    # 2. Header Structure
    headers = ["No.", "Nama Siswa", "Hadir", "Izin", "Sakit", "Alpha", "Total Hari", "Persentase Kehadiran (%)"]
    ws.append(headers)

    # Header Styling
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    header_font = Font(name='Calibri', size=12, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 3. Fill Data
    for i, row in df.iterrows():
        row_num = i + 2
        # No, Nama, Hadir, Izin, Sakit, Alpha
        ws.cell(row=row_num, column=1, value=i+1)
        ws.cell(row=row_num, column=2, value=row['Nama Siswa'])
        ws.cell(row=row_num, column=3, value=row['Hadir'])
        ws.cell(row=row_num, column=4, value=row['Izin'])
        ws.cell(row=row_num, column=5, value=row['Sakit'])
        ws.cell(row=row_num, column=6, value=row['Alpha'])
        
        # Formulas
        ws.cell(row=row_num, column=7, value=f"=SUM(C{row_num}:F{row_num})")
        # Formula for percentage: (Hadir / Total) * 100. Handle division by zero.
        ws.cell(row=row_num, column=8, value=f'=IF(G{row_num}>0, (C{row_num}/G{row_num}), 0)')
        ws.cell(row=row_num, column=8).number_format = '0.00%'

    last_row = len(df) + 1
    total_row = last_row + 1

    # 4. Formatting
    # Column Widths
    widths = {'A': 5, 'B': 30, 'C': 10, 'D': 10, 'E': 10, 'F': 10, 'G': 12, 'H': 25}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Alignment & Borders
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    data_font = Font(name='Calibri', size=11)

    for row in ws.iter_rows(min_row=2, max_row=total_row, min_col=1, max_col=8):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            if cell.column == 2:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

    # 5. Conditional Formatting for Column H
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    # H < 80% (0.8)
    ws.conditional_formatting.add(f'H2:H{last_row}',
        CellIsRule(operator='lessThan', formula=['0.8'], fill=red_fill))
    # H >= 90% (0.9)
    ws.conditional_formatting.add(f'H2:H{last_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=['0.9'], fill=green_fill))

    # 6. Total Row
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    for col_idx in range(3, 8):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{last_row})").font = Font(bold=True)
    
    # Percentage for Total Row requires a different formula to be accurate
    ws.cell(row=total_row, column=8, value=f'=IF(G{total_row}>0, (C{total_row}/G{total_row}), 0)')
    ws.cell(row=total_row, column=8).number_format = '0.00%'

    # 7. Pie Chart
    chart = PieChart()
    labels = Reference(ws, min_col=3, max_col=6, min_row=1)
    data = Reference(ws, min_col=3, max_col=6, min_row=total_row)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    chart.title = "Distribusi Kehadiran Keseluruhan"
    ws.add_chart(chart, "J2")

    # 8. Freeze Panes
    ws.freeze_panes = "A2"

    # Save
    wb.save(output_xlsx)
    print(f"File berhasil disimpan: {output_xlsx}")

if __name__ == "__main__":
    process_rekap("Rekap_Absensi_XII_IPA_2026-02-17.csv", "Rekap_Absensi_Rapi_XII_IPA.xlsx")
